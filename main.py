
from styles import *
from os import listdir
from openpyxl import Workbook
from string import ascii_uppercase
from enum import Enum
from time import sleep
import pdfplumber
# import PyPDF2

# TODO
# finish venue
# Clean up code
# Comment code
# Cross Check that code works
# Run on all PDfs

# Program requires format of files to be <CompanyName>(QX).pdf
# X is the quarter number

class State(Enum):
    """ State class to track which table the program is reading.
        Needed because some tables span across multiple pages.
        Idea is expanded more upon later.
        Tables always appear in this order. """

    SP_SUMMARY = 1
    SP_VENUE = 2
    NON_SP_SUMMARY = 3
    NON_SP_VENUE = 4
    OPTIONS_SUMMARY = 5
    OPTIONS_VENUE = 6


class SEC606Scraper:
    """ Main driver class """

    # Takes arguments for the names of the two outputted files
    def __init__(self, summaryFileName: str, venueFileName: str):

        self.pdfDirectory = "./UROPData"
        self.pdfs = listdir(self.pdfDirectory)  # Creates a list of all of the files (PDFs) in directory
        self.pdfs.sort()  # Sorts list so it is in alphabetical order (not sure if this is necessary but why not)

        if self.pdfs[0] == ".DS_Store":  # Delete .DS_Store Metadata file
            self.pdfs.pop(0)

        self.summaryFileName = summaryFileName
        self.venueFileName = venueFileName

        self.summaryFieldNames = ["Broker Name", "Year", "Quarter/Month", "Stock/Option"]
        self.venueFieldNames = ["Broker Name", "Year", "Quarter/Month", "Venue - Non-Directed Order Flow", "Stock/Option"]

        # Used to convert self.currentMonth to a string and self.state to a string (cleaner code)
        self.monthDict = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep",
                          10: "Oct", 11: "Nov", 12: "Dec"}
        self.stockOrOptionDict = {State.SP_SUMMARY: "S&P 500 Stocks", State.SP_VENUE: "S&P 500 Stocks",
                              State.NON_SP_SUMMARY: "Non-S&P 500 Stocks", State.NON_SP_VENUE: "Non-S&P 500 Stocks",
                              State.OPTIONS_SUMMARY: "Options", State.OPTIONS_VENUE: "Options"}

        # Important member variables
        self.currentCompanyName = ""
        self.makeNewSummarySheet = False
        self.makeNewVenueSheet = False
        self.currentQuarter = 1
        self.currentMonth = 1
        self.state = State.SP_SUMMARY  # First table is always Summary Table of the S&P 500
        self.YEAR = "2020"
        self.currentVenueRow = 2
        self.numSheets = 0

        # Included these companies from the start because I noticed their PDFs are structured differently
        # or they don't start from Q1
        self.unreadableFiles = ["WOOD (ARTHUR W.) COMPANY, INC. ", "VANDERBILT SECURITIES, LLC ", "G.RESEARCH, LLC ",
                                "CAPITAL INSTITUTIONAL SERVICES, INC. ", "FSC SECURITIES CORPORATION ",
                                "GILDER GAGNON HOWE _ CO. LLC ", "KEYBANC CAPITAL MARKETS INC. ",
                                "MESIROW FINANCIAL, INC. ", "MML INVESTORS SERVICES, LLC ", "MUTUAL SECURITIES, INC. ",
                                "PLANMEMBER SECURITIES CORPORATION ", "RAYMOND JAMES FINANCIAL SERVICES, INC. ",
                                "SAMUEL A. RAMIREZ _ COMPANY, INC. ", "WILLIAM BLAIR _ COMPANY L.L.C. "]

        # Make new Excel files and remove default Sheet
        self.venueFile = Workbook()
        self.summaryFile = Workbook()

        self.venueFile.remove(self.venueFile["Sheet"])
        self.summaryFile.remove(self.summaryFile["Sheet"])

    # Root of the function calls
    def getPDFs(self) -> None:
        # Loop through all of the PDFs
        for pdfIndex in range(len(self.pdfs)):
            pdfName = self.pdfs[pdfIndex]

            if self.currentCompanyName != "":
                assert self.state == State.OPTIONS_VENUE

            # If the PDF's name is not different from the one before it, we must be looking at the same company
            # but a new quarter
            if pdfName[0:pdfName.find('(')] == self.currentCompanyName:
                self.currentQuarter += 1

            # If not, it is a new company
            else:
                # Check if the new company has a format that's not readable or doesn't start at Q1
                if self.PDFNotReadable(pdfName):
                    print("PDF Index: " + str(pdfIndex) + " " + pdfName + " SKIPPED (Not Valid)")
                    continue

                # Else reset everything
                else:
                    if self.currentCompanyName != "":
                        assert self.currentQuarter == 4

                    self.makeNewSummarySheet = True
                    self.makeNewVenueSheet = True
                    self.currentQuarter = 1
                    self.currentMonth = 1
                    self.currentVenueRow = 2
                    self.state = State.SP_SUMMARY
                    self.currentCompanyName = pdfName[0:pdfName.find('(')]

            print("PDF Index: " + str(pdfIndex) + " " + pdfName)

            # Use pdfPlumber to open the PDF
            pdf = pdfplumber.open(self.pdfDirectory + '/' + pdfName)
            self.getPages(pdf)

            # sleep(1)

    # Loop through all of the PDF's pages
    def getPages(self, pdf) -> None:
        for pageIndex in range(len(pdf.pages)):
            page = pdf.pages[pageIndex]
            # print("Page Number: " + str(pageIndex))
            self.extractTableData(page)

    # Extract page's tables
    def extractTableData(self, page) -> None:
        tables = page.extract_tables()

        if len(tables) == 0:
            print("WARNING, TABLES LENGTH 0 ON PAGE " + str(page.page_number))

        # For each table
        for table in tables:
            fields = table[0]

            # self.state represents the previous page's last table in this section
            # The current table IS a summary table
            if "% of All Orders" in fields[0] or "All Orders" in fields[0]:
                self.newPageSummaryHelper()

            # The current table IS a venue table
            else:
                assert "Venue" in fields[0]
                self.newPageVenueHelper()

            # Write to venue or summary file depending on state
            if self.inSummaryState():
                # Is this right? Finish
                # Try to access the data of the table - if the table doesn't have any data (whether it's empty or
                # hanging across two pages), catch the error and loop to the next table
                try:
                    data = table[1]
                except IndexError:
                    print("Incomplete Summary Table Warning (only column titles, no data)     Page: " + str(page.page_number))
                    continue

                self.writeSummarySheet(data, fields)

            else:
                assert self.inVenueState()
                # Is this right? Finish
                try:
                    data = table[1:len(table)]  # 1:len(table) because Summary tables usually have > 1 lists inside them
                except IndexError:
                    print("Incomplete Venue Table Warning (only column titles, no data)     Page: " + str(page.page_number))
                    continue

                self.writeVenueSheet(data, fields[1:len(fields)])

    # Writes to summary file/sheet
    def writeSummarySheet(self, data: list, tableFields: list) -> None:

        sheet = self.createSheet(self.summaryFile, self.makeNewSummarySheet)

        # If a new sheet was made, format the sheet and write initial data to it
        if self.makeNewSummarySheet:
            self.makeNewSummarySheet = False
            self.increaseSummaryColumnWidth(sheet)
            self.writeHeaders(sheet, self.summaryFieldNames, tableFields)
            self.writeCell(sheet, 2, 1, self.currentCompanyName)
            self.writeCell(sheet, 2, 2, self.YEAR)

        # Decide which row to write to based on state (so the same categories are grouped together)
        if self.state == State.SP_SUMMARY:
            row = 2
        elif self.state == State.NON_SP_SUMMARY:
            row = 15
        else:
            assert self.state == State.OPTIONS_SUMMARY
            row = 28

        # Write quarter and type of stock/option trade
        row = row + self.currentMonth - 1

        self.writeCell(sheet, row, 3, "Q" + str(self.currentQuarter) + ", " + self.monthDict[self.currentMonth])
        self.writeCell(sheet, row, 4, self.stockOrOptionDict[self.state])

        # Loop through and write data underneath its respective header
        index = len(self.summaryFieldNames) + 1

        for value in data:
            self.writeCell(sheet, row, index, value)
            index += 1

    # Writes to venue file/sheet
    def writeVenueSheet(self, data: list, tableFields: list) -> None:

        sheet = self.createSheet(self.venueFile, self.makeNewVenueSheet)

        # If a new sheet was made, format the sheet and write initial data to it
        if self.makeNewVenueSheet:
            self.makeNewVenueSheet = False
            self.increaseVenueColumnWidth(sheet)
            self.writeHeaders(sheet, self.venueFieldNames, tableFields)
            self.writeCell(sheet, 2, 1, self.currentCompanyName)
            self.writeCell(sheet, 2, 2, self.YEAR)

        index = len(self.venueFieldNames) + 1

        # For each of the venues, first write the date, the venue name, and the security category
        for venueRow in data:
            row = self.currentVenueRow + self.currentMonth - 1
            self.writeCell(sheet, row, 3, "Q" + str(self.currentQuarter) + ", " + self.monthDict[self.currentMonth])
            self.writeCell(sheet, row, 4, venueRow[0])
            self.writeCell(sheet, row, 5, self.stockOrOptionDict[self.state])

            # Loop through each of the data in venueRow (excluding venueRow[0], which was the name), and write it
            for value in venueRow[1:len(venueRow)]:
                self.writeCell(sheet, row, index, value)
                index += 1

            # After a venue is written, increment the row but reset the column
            self.currentVenueRow += 1
            index = len(self.venueFieldNames) + 1

    # Creates a new sheet if makeSheet is True, if not returns the company's sheet that already exists
    def createSheet(self, file, makeSheet: bool):
        if makeSheet:
            sheet = file.create_sheet(self.currentCompanyName)
            self.numSheets += 1
        else:
            sheet = file[self.currentCompanyName]

        return sheet

    # Helper function for keeping track of state
    # Function runs when the current page's first table is a summary table
    def newPageSummaryHelper(self) -> None:
        # If the previous page's last table was a SP summary table and this one is a summary table too,
        # it must be the same
        if self.state == State.SP_SUMMARY:
            pass
        # If the previous page's last table was a SP venue table, then this must be Non SP summary
        elif self.state == State.SP_VENUE:
            self.incrementState()

        # Read the following conditionals the same as above
        # Pass means nothing should change - the state remains the same
        elif self.state == State.NON_SP_SUMMARY:
            pass
        elif self.state == State.NON_SP_VENUE:
            self.incrementState()

        elif self.state == State.OPTIONS_SUMMARY:
            pass
        # Incrementing state when you're at an Options Venue Table means the current month is over
        else:
            assert self.state == State.OPTIONS_VENUE
            self.incrementState()
            self.currentMonth += 1

    # Second helper function for keeping track of state
    # Function runs when the current page's first table is a venue table
    def newPageVenueHelper(self) -> None:
        # If the previous page's last table was a SP summary and you're on a venue now, then you must
        # be in a SP venue table
        if self.state == State.SP_SUMMARY:
            self.incrementState()
        # If the previous page's last table was a SP venue table and you're still in a venue, it must still
        # be SP venue
        elif self.state == State.SP_VENUE:
            pass

        elif self.state == State.NON_SP_SUMMARY:
            self.incrementState()
        elif self.state == State.NON_SP_VENUE:
            pass

        elif self.state == State.OPTIONS_SUMMARY:
            self.incrementState()
        else:
            assert self.state == State.OPTIONS_VENUE

    # Helper function for determining if a PDF is readable
    def PDFNotReadable(self, pdfName: str) -> bool:
        name = pdfName[0:pdfName.find('(')]

        if name in self.unreadableFiles:
            return True

        # File must end in .pdf
        elif pdfName[len(pdfName) - 4:len(pdfName)] != ".pdf":
            self.unreadableFiles.append(name)
            print("Doesn't end in PDF")
            return True

        # If file doesn't have Q1
        elif pdfName[pdfName.find('(') + 1:pdfName.find('(') + 3] != "Q1":
            self.unreadableFiles.append(name)
            print("Doesn't start with Q1")
            return True

        elif name + "(Q2).pdf" not in self.pdfs or \
             name + "(Q3).pdf" not in self.pdfs or \
             name + "(Q4).pdf" not in self.pdfs:

            self.unreadableFiles.append(name)
            print("Doesn't have all four Quarters")
            return True

        return False

    # Helper function for incrementing state
    def incrementState(self):
        if self.state == State.OPTIONS_VENUE:  # Wrap state back to beginning
            self.state = State.SP_SUMMARY
        else:
            self.state = State(self.state.value + 1)

        # print("Incremented to: " + str(self.state.name))

    # Helper function for writing the first row of new sheets
    def writeHeaders(self, sheet, fields: list, additionalFields: list) -> None:
        index = 1
        finalFields = fields + additionalFields  # Combine the above lists together

        for field in finalFields:
            self.writeCell(sheet, 1, index, field)
            index += 1

    # Helper function for writing to cells
    def writeCell(self, sheet, row: int, col: int, value) -> None:
        cell = sheet.cell(row, col)
        cell.value = value
        cell.alignment = CENTER

    # Both helper functions below increase the columns' width to the right size
    def increaseSummaryColumnWidth(self, sheet) -> None:
        sheet.column_dimensions['A'].width = 40

        for col in ascii_uppercase[1:4]:
            sheet.column_dimensions[col].width = 18

        for col in ascii_uppercase[4:10]:
            sheet.column_dimensions[col].width = 50

    def increaseVenueColumnWidth(self, sheet) -> None:
        sheet.column_dimensions['A'].width = 40

        for col in ascii_uppercase[1:3]:
            sheet.column_dimensions[col].width = 18

        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 20

        for col in ascii_uppercase[5:10]:
            sheet.column_dimensions[col].width = 40

        sheet.column_dimensions['K'].width = 65
        sheet.column_dimensions['L'].width = 75
        sheet.column_dimensions['M'].width = 60
        sheet.column_dimensions['N'].width = 65
        sheet.column_dimensions['O'].width = 60
        sheet.column_dimensions['P'].width = 65
        sheet.column_dimensions['Q'].width = 60
        sheet.column_dimensions['R'].width = 60

    # Both helper functions below return whether or not we're in summary or venue state
    def inSummaryState(self) -> bool:
        return self.state == State.SP_SUMMARY or self.state == State.NON_SP_SUMMARY or self.state == State.OPTIONS_SUMMARY

    def inVenueState(self) -> bool:
        return self.state == State.SP_VENUE or self.state == State.NON_SP_VENUE or self.state == State.OPTIONS_VENUE

    # Save Excel files
    def save(self) -> None:
        self.summaryFile.save(self.summaryFileName)
        self.venueFile.save(self.venueFileName)

    # Write unreadable files to a txt file so the user knows
    def outputUnreadableFiles(self) -> None:
        file = open("unreadable files.txt", 'w')
        self.unreadableFiles.sort()

        for filename in self.unreadableFiles:
            file.write(filename + '\n')

        file.close()

        print("UNREADABLE FILES:")
        print(self.unreadableFiles)


# Run SEC606Scraper!
def main():
    scraper = SEC606Scraper("summary.xlsx", "venue.xlsx")
    scraper.getPDFs()
    scraper.save()
    scraper.outputUnreadableFiles()

    print()
    print("Number of company's data gathered: " + str(scraper.numSheets/2))
    print("Number of company's data ignored: " + str(len(scraper.unreadableFiles)))


main()
